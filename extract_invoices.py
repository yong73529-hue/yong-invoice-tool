from __future__ import annotations

import re
import os
import sys
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
import pdfplumber
import pypdfium2 as pdfium
import numpy as np
from PIL import Image
from openpyxl.styles import PatternFill

try:
    from rapidocr_onnxruntime import RapidOCR
    RAPIDOCR_IMPORT_ERROR = ""
except Exception as exc:
    RapidOCR = None
    RAPIDOCR_IMPORT_ERROR = str(exc)


def resolve_base_dir() -> Path:
    env_base_dir = os.environ.get("YONG_INVOICE_BASE_DIR")
    if env_base_dir:
        return Path(env_base_dir).expanduser().resolve()

    return Path.cwd().resolve()


BASE_DIR = resolve_base_dir()
INPUT_DIR = BASE_DIR / "invoices"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "invoice_summary.xlsx"
ACTIVE_OUTPUT_FILE = OUTPUT_FILE
SUPPORTED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}
OCR_ENGINE = None

COLUMNS = [
    "录入日期",
    "文件名",
    "发票号码",
    "开票日期",
    "发票类型",
    "金额",
    "税额",
    "价税合计",
    "项目名称",
    "销售方名称",
    "销售方纳税人识别号",
    "购买方名称",
    "购买方纳税人识别号",
    "处理状态",
    "备注",
]


def normalize_text(text: str) -> str:
    text = text.replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def compact_text(text: str) -> str:
    return re.sub(r"\s+", "", text)


def read_pdf_text(pdf_path: Path) -> str:
    parts: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                parts.append(page_text)
    return normalize_text("\n".join(parts))


def read_document_text(file_path: Path) -> tuple[str, bool]:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        text = read_pdf_text(file_path)
        if text:
            return text, False
        return read_pdf_ocr_text(file_path), True

    if suffix in SUPPORTED_EXTENSIONS:
        return read_image_ocr_text(file_path), True

    raise ValueError(f"不支持的文件类型：{suffix}")


def read_pdf_ocr_text(pdf_path: Path) -> str:
    return normalize_ocr_text("\n".join(read_pdf_ocr_pages(pdf_path)))


def read_pdf_ocr_pages(pdf_path: Path) -> list[str]:
    engine = get_ocr_engine()
    parts: list[str] = []
    pdf = pdfium.PdfDocument(str(pdf_path))
    try:
        for page in pdf:
            image = page.render(scale=3).to_pil()
            page_text = run_ocr(engine, image)
            if page_text:
                parts.append(normalize_ocr_text(page_text))
    finally:
        pdf.close()
    return parts


def read_image_ocr_text(image_path: Path) -> str:
    engine = get_ocr_engine()
    with Image.open(image_path) as image:
        return normalize_ocr_text(run_ocr(engine, image.convert("RGB")))


def get_ocr_engine():
    global OCR_ENGINE
    if RapidOCR is None:
        detail = f"；导入错误：{RAPIDOCR_IMPORT_ERROR}" if RAPIDOCR_IMPORT_ERROR else ""
        raise RuntimeError(f"缺少OCR依赖，请先运行 install_ocr_dependencies.bat，装完后关闭Excel并重新运行 run.bat{detail}")
    if OCR_ENGINE is None:
        OCR_ENGINE = RapidOCR()
    return OCR_ENGINE


def run_ocr(engine, image: Image.Image) -> str:
    result, _ = engine(np.array(image))
    if not result:
        return ""

    lines = []
    for item in result:
        if len(item) >= 2:
            lines.append(str(item[1]).strip())
    return "\n".join(line for line in lines if line)


def normalize_ocr_text(text: str) -> str:
    text = normalize_text(text)
    text = re.sub(r"([0-9])\s*\.\s*([0-9]{1,2})", r"\1.\2", text)
    text = re.sub(r"([¥￥´`])\s+([0-9])", r"\1\2", text)
    return text


def first_match(patterns: list[str], text: str) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            value = match.group(1).strip()
            value = re.sub(r"^[：:\s]+|[：:\s]+$", "", value)
            if value:
                return value
    return ""


def extract_invoice_type(text: str, compact: str) -> str:
    if "铁路电子客票" in compact or ("铁路" in compact and "电子客票" in compact):
        return "铁路电子客票"

    value = first_match(
        [
            r"电子发票[（(]\s*([^）)\n]+?)\s*[）)]",
            r"(增值税专用发票|普通发票)",
            r"([普曾]通发票|普发票)",
        ],
        text,
    )
    if value:
        if value in {"曾通发票", "普发票"}:
            return "普通发票"
        return value

    match = re.search(r"电子发票[（(]([^）)]+)[）)]", compact)
    if match:
        return match.group(1)
    if "增值税专用发票" in compact:
        return "增值税专用发票"
    if "增值税" in compact and "用发票" in compact:
        return "增值税专用发票"
    if "用发票" in compact and "普通发票" not in compact and "普发票" not in compact:
        return "增值税专用发票"
    if "普通发票" in compact or "曾通发票" in compact or "普发票" in compact:
        return "普通发票"
    return ""


def extract_invoice_number(text: str, compact: str) -> str:
    value = first_match(
        [
            r"发票号码\s*[：:]\s*([0-9]{8,30})",
            r"发票号\s*码?\s*[：:]\s*([0-9]{8,30})",
            r"Invoice\s*No\.?\s*[：:]?\s*([0-9]{8,30})",
        ],
        text,
    )
    if value:
        return value

    match = re.search(r"发票号码[:：]?([0-9]{8,30})", compact)
    if match:
        return match.group(1)

    for line in text.splitlines():
        line_compact = compact_text(line)
        if re.fullmatch(r"[0-9]{20}", line_compact):
            return line_compact

    candidates = re.findall(r"(?<![A-Z0-9])([0-9]{20})(?![A-Z0-9])", compact)
    return candidates[0] if candidates else ""


def extract_invoice_date(text: str, compact: str) -> str:
    value = first_match(
        [
            r"开票日期\s*[：:]\s*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
            r"开票日期\s*[：:]\s*([0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2})",
            r"填开日期\s*[：:]\s*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
        ],
        text,
    )
    if not value:
        match = re.search(
            r"(?:开票日期|填开日期)[:：]?([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日|[0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2})",
            compact,
        )
        value = match.group(1) if match else ""
    if not value:
        for line in text.splitlines():
            line_compact = compact_text(line)
            if re.fullmatch(r"[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日", line_compact):
                value = line_compact
                break
    if not value:
        match = re.search(r"([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)", compact)
        value = match.group(1) if match else ""
    return normalize_date(value)


def normalize_date(value: str) -> str:
    if not value:
        return ""
    match = re.match(r"([0-9]{4})年([0-9]{1,2})月([0-9]{1,2})日", value)
    if match:
        year, month, day = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    match = re.match(r"([0-9]{4})[-/.]([0-9]{1,2})[-/.]([0-9]{1,2})", value)
    if match:
        year, month, day = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return value


def extract_amount(text: str, compact: str) -> str:
    for line in text.splitlines():
        compact_line = compact_text(line)
        if "合计" in compact_line and "价税合计" not in compact_line:
            numbers = extract_money_numbers(line)
            if numbers:
                return normalize_amount(numbers[0])

    patterns = [
        r"金额\s*[：:]?\s*[¥￥]?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
    ]
    value = first_match(patterns, text)
    if not value:
        match = re.search(
            r"金额[:：]?[¥￥]?([0-9,]+\.[0-9]{2})",
            compact,
        )
        value = match.group(1) if match else ""
    if not value:
        pair = extract_last_currency_pair(text)
        value = pair[0] if pair else ""
    if not value:
        amounts = extract_currency_amounts(text)
        value = amounts[0] if amounts else ""
    return normalize_amount(value)


def extract_tax_amount(text: str, compact: str) -> str:
    for line in text.splitlines():
        compact_line = compact_text(line)
        if "合计" in compact_line and "价税合计" not in compact_line:
            numbers = extract_money_numbers(line)
            if len(numbers) >= 2:
                return normalize_amount(numbers[-1])

    value = first_match(
        [
            r"税额\s*[：:]?\s*[¥￥]?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
        ],
        text,
    )
    if not value:
        match = re.search(r"税额[:：]?[¥￥]?([0-9,]+\.[0-9]{2})", compact)
        value = match.group(1) if match else ""
    if not value:
        pair = extract_last_currency_pair(text)
        value = pair[1] if pair else ""
    if not value:
        amounts = extract_currency_amounts(text)
        value = amounts[1] if len(amounts) >= 2 else ""
    if not value:
        value = extract_tax_amount_after_rate(text)
    return normalize_amount(value)


def extract_total_amount(text: str, compact: str) -> str:
    patterns = [
        r"(?:价税合计|小写)\s*[）)]?\s*[：:]?\s*[¥￥´`]?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
        r"价税合计.*?小写\s*[）)]?\s*[：:]?\s*[¥￥´`]?\s*([0-9,]+(?:\.[0-9]{1,2})?)",
    ]
    value = first_match(patterns, text)
    if not value:
        match = re.search(
            r"(?:价税合计|小写)[:：]?[¥￥´`]?([0-9,]+\.[0-9]{2})",
            compact,
        )
        value = match.group(1) if match else ""
    if not value:
        matches = extract_currency_amounts(text)
        value = matches[-1] if matches else ""
    return normalize_amount(value)


def extract_money_numbers(value: str) -> list[str]:
    return re.findall(r"[¥￥]?\s*([0-9,]+\.[0-9]{1,2})", value)


def extract_last_currency_pair(text: str) -> tuple[str, str] | None:
    pairs: list[tuple[str, str]] = []
    for line in text.splitlines():
        numbers = re.findall(r"[¥￥]\s*([0-9,]+\.[0-9]{1,2})", line)
        if len(numbers) >= 2:
            pairs.append((numbers[0], numbers[1]))
    return pairs[-1] if pairs else None


def extract_currency_amounts(text: str) -> list[str]:
    return re.findall(r"[¥￥´`]\s*([0-9,]+\.[0-9]{1,2})", text)


def extract_tax_amount_after_rate(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        if "%" in line:
            for next_line in lines[index + 1 : index + 4]:
                match = re.search(r"([0-9,]+\.[0-9]{1,2})", next_line)
                if match:
                    return match.group(1)
    return ""


def normalize_amount(value: str) -> str:
    if not value:
        return ""
    try:
        return f"{Decimal(value.replace(',', '')):.2f}"
    except InvalidOperation:
        return value


def calculate_tax_amount(amount: str, total_amount: str) -> str:
    if not amount or not total_amount:
        return ""
    try:
        tax = Decimal(total_amount.replace(",", "")) - Decimal(amount.replace(",", ""))
        if tax >= 0:
            return f"{tax:.2f}"
    except InvalidOperation:
        return ""
    return ""


def calculate_total_amount(amount: str, tax_amount: str) -> str:
    if not amount or not tax_amount:
        return ""
    try:
        total = Decimal(amount.replace(",", "")) + Decimal(tax_amount.replace(",", ""))
        return f"{total:.2f}"
    except InvalidOperation:
        return ""


def extract_party_info(text: str, compact: str, role: str) -> str:
    split_info = extract_split_party_info(text, role)
    if split_info:
        return split_info

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    other_role = "销售方" if role == "购买方" else "购买方"
    stop_markers = [other_role, "货物或应税劳务", "项目名称", "合计", "价税合计", "备注", "收款人"]

    for index, line in enumerate(lines):
        if role in line:
            parts = [clean_party_info_line(line.replace(role, ""))]
            for next_line in lines[index + 1 : index + 10]:
                if any(marker in next_line for marker in stop_markers):
                    break
                cleaned = clean_party_info_line(next_line)
                if cleaned:
                    parts.append(cleaned)
            value = "；".join(part for part in parts if part)
            if value:
                return value

    stop_pattern = "销售方|货物或应税劳务|项目名称|合计|价税合计|备注|收款人" if role == "购买方" else "购买方|备注|收款人|复核|开票人"
    match = re.search(rf"{role}(.+?)(?:{stop_pattern})", compact)
    if match:
        return clean_party_info_line(match.group(1))
    return ""


def extract_party_fields(text: str, compact: str, role: str) -> tuple[str, str]:
    info = extract_party_info(text, compact, role)
    return parse_party_info(info)


def parse_party_info(info: str) -> tuple[str, str]:
    if not info:
        return "", ""

    name = first_match(
        [
            r"名\s*称\s*[：:]\s*(.*?)(?=\s*(?:统一社会信用代码|统一社会信息代码|统一社会信用代码/纳税人识别号|统一社会信息代码/纳税人识别号|纳税人识别号)[：:]|[；;\n]|$)",
        ],
        info,
    )
    tax_id = first_match(
        [
            r"(?:统一社会信用代码/纳税人识别号|统一社会信息代码/纳税人识别号|纳税人识别号)\s*[：:]\s*([A-Z0-9]+)",
            r"(?:统一社会信用代码|统一社会信息代码)\s*[：:]\s*([A-Z0-9]+)",
        ],
        info,
    )
    return name, tax_id


def extract_split_party_info(text: str, role: str) -> str:
    block_info = extract_party_info_from_ocr_blocks(text, role)
    if block_info:
        return block_info

    name_match = re.search(
        r"购\s*名\s*称\s*[：:]?\s*(.*?)\s+销\s*名\s*称\s*[：:]?\s*([^\n]+)",
        text,
    )
    if not name_match:
        name_match = re.search(
            r"买\s*名\s*称\s*[：:]?\s*(.*?)\s+售\s*名\s*称\s*[：:]?\s*([^\n]+)",
            text,
        )
    if not name_match:
        name_match = re.search(
            r"名\s*称\s*[：:]?\s*(.*?)\s+名\s*称\s*[：:]?\s*([^\n]+)",
            text,
        )
    tax_ids = re.findall(r"统一社会(?:信用|信息)代码/?纳税人识别号[：:]?([A-Z0-9]+)", compact_text(text))
    if not tax_ids:
        tax_ids = extract_split_tax_ids_from_lines(text)
    fallback_names = extract_split_party_names_from_lines(text)

    role_index = 0 if role == "购买方" else 1
    parts: list[str] = []
    if name_match:
        name = clean_party_info_line(name_match.group(role_index + 1))
        if is_valid_party_name(name):
            parts.append(f"名称：{name}")
    if not parts and len(fallback_names) > role_index:
        parts.append(f"名称：{fallback_names[role_index]}")
    if len(tax_ids) > role_index:
        parts.append(f"统一社会信用代码/纳税人识别号：{tax_ids[role_index]}")

    return "；".join(parts)


def extract_party_info_from_ocr_blocks(text: str, role: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    names: dict[str, str] = {}
    for index, line in enumerate(lines):
        if line in {"购买方信息", "销售方信息"}:
            role_name = "购买方" if line == "购买方信息" else "销售方"
            for next_line in lines[index + 1 : index + 10]:
                match = re.search(r"名称\s*[：:]?\s*(.+)", next_line)
                if match:
                    name = clean_party_info_line(match.group(1))
                    if is_valid_party_name(name):
                        names[role_name] = name
                        break
                cleaned = clean_party_info_line(next_line)
                if is_valid_party_name(cleaned):
                    names[role_name] = cleaned
                    break

    if role not in names:
        return ""

    tax_ids = re.findall(r"统一社会(?:信用|信息)代码/?纳税人识别号[：:]?([A-Z0-9]+)", compact_text(text))
    if len(tax_ids) < 2:
        tax_ids = extract_tax_ids_anywhere(text)
    role_index = 0 if role == "购买方" else 1
    parts = [f"名称：{names[role]}"]
    if len(tax_ids) > role_index:
        parts.append(f"统一社会信用代码/纳税人识别号：{tax_ids[role_index]}")
    return "；".join(parts)


def extract_tax_ids_anywhere(text: str) -> list[str]:
    candidates = re.findall(r"(?<![A-Z0-9])(?=[A-Z0-9]*[A-Z])[A-Z0-9]{15,20}(?![A-Z0-9])", text)
    return list(dict.fromkeys(candidates))


def extract_split_tax_ids_from_lines(text: str) -> list[str]:
    for line in text.splitlines():
        tax_ids = re.findall(r"(?<![A-Z0-9])[A-Z0-9]{15,20}(?![A-Z0-9])", line)
        if len(tax_ids) >= 2:
            return tax_ids[:2]
    return []


def is_valid_party_name(value: str) -> bool:
    if len(value) < 4:
        return False
    if re.fullmatch(r"[购买销售方信信息息名称：:]+", value):
        return False
    return any(
        marker in value
        for marker in [
            "公司",
            "店",
            "中心",
            "商行",
            "门市部",
            "工作室",
            "经营部",
            "制作部",
            "服务部",
            "食铺",
            "饭店",
            "餐馆",
            "餐饮",
            "小吃",
            "客栈",
            "客栈坊",
            "宾馆",
            "酒店",
            "旅馆",
            "超市",
            "花店",
            "贸易",
            "科技",
        ]
    )


def extract_split_party_names_from_lines(text: str) -> list[str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    tax_line_index = -1
    for index, line in enumerate(lines):
        if len(re.findall(r"(?<![A-Z0-9])[A-Z0-9]{15,20}(?![A-Z0-9])", line)) >= 2:
            tax_line_index = index
            break

    if tax_line_index <= 0:
        return []

    line = lines[tax_line_index - 1]
    company_pattern = r".+?(?:有限公司|分公司|个体工商户|中心|商行|门市部|工作室|经营部|制作部|服务部|食铺|饭店|餐馆|餐饮店|小吃店|客栈坊|客栈|宾馆|酒店|旅馆|超市|花店|店)"
    match = re.match(rf"({company_pattern})\s+({company_pattern}.*)", line)
    if match:
        return [clean_party_info_line(match.group(1)), clean_party_info_line(match.group(2))]
    return []


def clean_party_info_line(value: str) -> str:
    value = re.sub(r"\s+", "", value)
    value = value.strip("：:；;")
    value = re.sub(r"[购买销售方信]+$", "", value)
    return value


def extract_item_name(text: str) -> str:
    compact = compact_text(text)
    if "铁路电子客票" in compact or ("铁路" in compact and "电子客票" in compact):
        return "铁路旅客运输服务"

    starred_items = re.findall(r"\*[^*\s\n]+\*[^\s\n]+", text)
    if starred_items:
        return "；".join(dict.fromkeys(starred_items))

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    header_index = -1
    for index, line in enumerate(lines):
        if ("货物" in line and ("劳务" in line or "服务" in line) and "名称" in line) or "项目名称" in line:
            header_index = index
            break

    if header_index >= 0:
        candidates: list[str] = []
        for line in lines[header_index + 1 : header_index + 8]:
            if any(marker in line for marker in ["合计", "价税合计", "销售方", "购买方", "备注"]):
                break
            item = line.split()[0].strip()
            if item and not re.fullmatch(r"[0-9,¥￥.\-*%]+", item):
                candidates.append(item)
        if candidates:
            return "；".join(dict.fromkeys(candidates))

    value = first_match(
        [
            r"(\*[^*\n]+\*[^\s\n]+)",
            r"项目名称\s*[：:]\s*([^\n]+)",
            r"商品名称\s*[：:]\s*([^\n]+)",
            r"服务名称\s*[：:]\s*([^\n]+)",
        ],
        text,
    )
    return value.strip()


def fill_railway_ticket_fields(row: dict[str, str], text: str, compact: str) -> None:
    if "铁路电子客票" not in compact and not ("铁路" in compact and "电子客票" in compact):
        return

    row["发票类型"] = "铁路电子客票"
    if not row["项目名称"]:
        row["项目名称"] = "铁路旅客运输服务"
    if not row["销售方名称"]:
        row["销售方名称"] = "中国国家铁路集团有限公司"
    if not row["销售方纳税人识别号"]:
        row["销售方纳税人识别号"] = "91100000000013477B"

    buyer_match = re.search(
        r"购买方名称\s*[：:]\s*(.*?)\s*统一社会信用代码\s*[：:]\s*([A-Z0-9]+)",
        text,
    )
    if not buyer_match:
        buyer_match = re.search(
            r"购买方名称[:：]?(.*?)统一社会信用代码[:：]?([A-Z0-9]+)",
            compact,
        )
    if buyer_match:
        row["购买方名称"] = clean_party_info_line(buyer_match.group(1))
        row["购买方纳税人识别号"] = buyer_match.group(2)


def parse_invoice(pdf_path: Path) -> dict[str, str]:
    row = {column: "" for column in COLUMNS}
    row["录入日期"] = datetime.fromtimestamp(pdf_path.stat().st_ctime).strftime("%Y-%m-%d")
    row["文件名"] = pdf_path.name

    try:
        text, used_ocr = read_document_text(pdf_path)
    except Exception as exc:
        row["处理状态"] = "失败"
        row["备注"] = f"文件读取失败：{exc}"
        return row

    if not text:
        row["处理状态"] = "失败"
        row["备注"] = "未提取到文本，OCR未识别到有效文字"
        return row

    return parse_invoice_text(row, text, used_ocr)


def parse_invoice_text(row: dict[str, str], text: str, used_ocr: bool) -> dict[str, str]:
    remarks: list[str] = []
    compact = compact_text(text)
    row["发票号码"] = extract_invoice_number(text, compact)
    row["开票日期"] = extract_invoice_date(text, compact)
    row["发票类型"] = extract_invoice_type(text, compact)
    row["金额"] = extract_amount(text, compact)
    row["税额"] = extract_tax_amount(text, compact)
    row["价税合计"] = extract_total_amount(text, compact)
    calculated_tax = calculate_tax_amount(row["金额"], row["价税合计"])
    if calculated_tax and (not row["税额"] or row["税额"] == row["价税合计"]):
        row["税额"] = calculated_tax
    calculated_total = calculate_total_amount(row["金额"], row["税额"])
    if calculated_total and (not row["价税合计"] or row["价税合计"] == row["金额"]):
        row["价税合计"] = calculated_total
    row["项目名称"] = extract_item_name(text)
    row["销售方名称"], row["销售方纳税人识别号"] = extract_party_fields(text, compact, "销售方")
    row["购买方名称"], row["购买方纳税人识别号"] = extract_party_fields(text, compact, "购买方")
    fill_railway_ticket_fields(row, text, compact)

    missing = [
        column
        for column in [
            "发票号码",
            "开票日期",
            "发票类型",
            "金额",
            "税额",
            "价税合计",
            "项目名称",
            "销售方名称",
            "销售方纳税人识别号",
            "购买方名称",
            "购买方纳税人识别号",
        ]
        if not row[column]
    ]
    if missing:
        row["处理状态"] = "部分成功"
        remarks.append("未识别字段：" + "、".join(missing))
    else:
        row["处理状态"] = "成功"
    if used_ocr:
        remarks.append("OCR识别，请人工复核")

    row["备注"] = "；".join(remarks)
    return row


def parse_invoice_rows(file_path: Path) -> list[dict[str, str]]:
    suffix = file_path.suffix.lower()
    base_row = {column: "" for column in COLUMNS}
    base_row["录入日期"] = datetime.fromtimestamp(file_path.stat().st_ctime).strftime("%Y-%m-%d")
    base_row["文件名"] = file_path.name

    if suffix == ".pdf":
        try:
            text = read_pdf_text(file_path)
        except Exception as exc:
            row = base_row.copy()
            row["处理状态"] = "失败"
            row["备注"] = f"文件读取失败：{exc}"
            return [row]

        if text:
            return [parse_invoice_text(base_row.copy(), text, False)]

        try:
            pages = read_pdf_ocr_pages(file_path)
        except Exception as exc:
            row = base_row.copy()
            row["处理状态"] = "失败"
            row["备注"] = f"OCR读取失败：{exc}"
            return [row]

        if not pages:
            row = base_row.copy()
            row["处理状态"] = "失败"
            row["备注"] = "未提取到文本，OCR未识别到有效文字"
            return [row]

        rows = []
        for index, page_text in enumerate(pages, start=1):
            row = base_row.copy()
            if len(pages) > 1:
                row["文件名"] = f"{file_path.name} 第{index}页"
            rows.append(parse_invoice_text(row, page_text, True))
        return rows

    return [parse_invoice(file_path)]


def read_existing_entry_dates() -> dict[str, str]:
    if not OUTPUT_FILE.exists():
        return {}

    try:
        old_df = pd.read_excel(OUTPUT_FILE, sheet_name="发票汇总", dtype=str).fillna("")
    except Exception as exc:
        print(f"未能读取旧汇总表，录入日期将按文件日期填写：{exc}")
        return {}

    if "文件名" not in old_df.columns or "录入日期" not in old_df.columns:
        return {}

    entry_dates: dict[str, str] = {}
    for _, old_row in old_df.iterrows():
        filename = str(old_row.get("文件名", "")).strip()
        entry_date = str(old_row.get("录入日期", "")).strip()
        if filename and entry_date and filename not in entry_dates:
            entry_dates[filename] = entry_date
    return entry_dates


def preserve_existing_entry_dates(rows: list[dict[str, str]], entry_dates: dict[str, str]) -> None:
    if not entry_dates:
        return

    kept_count = 0
    for row in rows:
        old_entry_date = entry_dates.get(row.get("文件名", ""))
        if old_entry_date:
            row["录入日期"] = old_entry_date
            kept_count += 1

    if kept_count:
        print(f"已保留旧汇总表中的录入日期：{kept_count} 条")


def mark_duplicate_invoice_numbers(rows: list[dict[str, str]]) -> None:
    counts = Counter(row["发票号码"] for row in rows if row.get("发票号码"))
    for row in rows:
        invoice_number = row.get("发票号码", "")
        if invoice_number and counts[invoice_number] > 1:
            duplicate_note = "发票号码重复"
            row["备注"] = f"{row['备注']}；{duplicate_note}" if row["备注"] else duplicate_note


def write_excel(rows: list[dict[str, str]]) -> None:
    global ACTIVE_OUTPUT_FILE
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=COLUMNS)
    df = df.sort_values(by=["录入日期", "文件名"], ascending=[True, True], kind="stable")
    output_file = OUTPUT_FILE
    try:
        writer = pd.ExcelWriter(output_file, engine="openpyxl")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"invoice_summary_{timestamp}.xlsx"
        print(f"默认Excel文件正在被占用，改为生成：{output_file}")
        writer = pd.ExcelWriter(output_file, engine="openpyxl")

    ACTIVE_OUTPUT_FILE = output_file
    with writer:
        df.to_excel(writer, index=False, sheet_name="发票汇总")
        worksheet = writer.sheets["发票汇总"]
        highlight_duplicate_invoice_numbers(worksheet, df)
        widths = {
            "A": 14,
            "B": 34,
            "C": 20,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 34,
            "J": 34,
            "K": 24,
            "L": 34,
            "M": 24,
            "N": 12,
            "O": 46,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        worksheet.freeze_panes = "A2"


def highlight_duplicate_invoice_numbers(worksheet, df: pd.DataFrame) -> None:
    invoice_number_column = COLUMNS.index("发票号码") + 1
    duplicated = df["发票号码"].astype(str).str.strip()
    duplicate_mask = duplicated.ne("") & duplicated.duplicated(keep=False)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    for row_index, is_duplicate in enumerate(duplicate_mask, start=2):
        if is_duplicate:
            worksheet.cell(row=row_index, column=invoice_number_column).fill = yellow_fill


def main() -> int:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    pdf_files = sorted(path for path in INPUT_DIR.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS)

    print(f"输入文件夹：{INPUT_DIR}")
    print(f"输出文件：{OUTPUT_FILE}")

    if not pdf_files:
        print("没有找到可处理文件。请把电子发票PDF或图片放入 invoices 文件夹后再运行。")
        return 0

    existing_entry_dates = read_existing_entry_dates()
    rows = []
    for pdf_path in pdf_files:
        print(f"正在处理：{pdf_path.name}")
        rows.extend(parse_invoice_rows(pdf_path))

    preserve_existing_entry_dates(rows, existing_entry_dates)
    mark_duplicate_invoice_numbers(rows)
    write_excel(rows)

    success_count = sum(1 for row in rows if row["处理状态"] == "成功")
    review_count = len(rows) - success_count
    print(f"处理完成：共 {len(rows)} 个PDF，成功 {success_count} 个，需复核 {review_count} 个。")
    print(f"Excel已生成：{ACTIVE_OUTPUT_FILE}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
